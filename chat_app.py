import streamlit as st
import pandas as pd
import joblib
import os
import io
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
from transformers import pipeline
from dotenv import load_dotenv

try:
    import plotly.express as px
    import plotly.graph_objects as go
    PLOTLY = True
except ImportError:
    PLOTLY = False

from langchain_google_genai import ChatGoogleGenerativeAI
from langchain.agents import AgentExecutor, create_tool_calling_agent
from langchain.tools import tool
from langchain.memory import ConversationBufferMemory
from langchain_core.prompts import ChatPromptTemplate

# ─────────────────────────────────────────────────────────────
# PAGE SETUP
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    layout="wide",
    page_title="NexaHR | Workforce Intelligence",
    page_icon="🏢",
    initial_sidebar_state="expanded"
)
load_dotenv()

# ─────────────────────────────────────────────────────────────
# DESIGN SYSTEM — CSS
# ─────────────────────────────────────────────────────────────
CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700;800&family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;1,9..40,400&family=DM+Mono:wght@400;500&display=swap');

/* ── RESET & BASE ── */
*, *::before, *::after { box-sizing: border-box; }
html, body, [data-testid="stAppViewContainer"] {
    font-family: 'DM Sans', sans-serif !important;
    background: #eef1f8 !important;
}
#MainMenu, footer { visibility: hidden; }
[data-testid="stToolbar"] { display: none !important; }
[data-testid="stDecoration"] { display: none !important; }

/* ── DARK SIDEBAR ── */
[data-testid="stSidebar"] {
    background: #0b1120 !important;
    border-right: 1px solid rgba(255,255,255,0.06) !important;
}

/* Force the Re-open arrow (>) to stay visible and clickable */
[data-testid="collapsedControl"] {
    z-index: 999999 !important;
}
}
}
[data-testid="stSidebar"] .stMarkdown p,
[data-testid="stSidebar"] .stMarkdown span {
    color: #94a3b8 !important;
    font-size: 0.85rem !important;
}
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3,
[data-testid="stSidebar"] h4 {
    color: #f1f5f9 !important;
    font-family: 'Sora', sans-serif !important;
}
[data-testid="stSidebar"] hr {
    border-color: #1e293b !important;
    opacity: 1 !important;
    margin: 0.75rem 0 !important;
}
[data-testid="stSidebar"] label { color: #94a3b8 !important; }
[data-testid="stSidebar"] button {
    background: #1a2540 !important;
    color: #94a3b8 !important;
    border: 1px solid #2d3f5c !important;
    border-radius: 8px !important;
    font-family: 'DM Sans', sans-serif !important;
    transition: all 0.2s !important;
}
[data-testid="stSidebar"] button:hover {
    background: #253556 !important;
    color: #e2e8f0 !important;
    border-color: #4b6cb7 !important;
}
[data-testid="stSidebar"] [data-testid="stSelectbox"] > div > div {
    background: #1a2540 !important;
    border-color: #2d3f5c !important;
    color: #e2e8f0 !important;
}
[data-testid="stSidebar"] .stSlider [data-baseweb="slider"] {
    margin-top: 0.5rem !important;
}
[data-testid="stSidebar"] details {
    border-color: #1e293b !important;
    border-radius: 10px !important;
    overflow: hidden !important;
}
[data-testid="stSidebar"] summary {
    background: #1a2540 !important;
    color: #94a3b8 !important;
    padding: 0.65rem 1rem !important;
    font-size: 0.85rem !important;
}

/* ── MAIN CONTENT ── */
.main .block-container {
    max-width: 1440px !important;
    padding: 2rem 2.5rem 3rem !important;
}

/* ── TYPOGRAPHY ── */
h1 {
    font-family: 'Sora', sans-serif !important;
    font-weight: 700 !important;
    color: #0f172a !important;
    letter-spacing: -0.5px !important;
    font-size: 1.75rem !important;
    line-height: 1.2 !important;
    margin-bottom: 0 !important;
}
h2 {
    font-family: 'Sora', sans-serif !important;
    font-weight: 600 !important;
    color: #1e293b !important;
    font-size: 1.25rem !important;
}
h3 {
    font-family: 'Sora', sans-serif !important;
    font-weight: 600 !important;
    color: #334155 !important;
    font-size: 1rem !important;
}
p, li { color: #475569 !important; }
strong { color: #1e293b !important; }
code {
    font-family: 'DM Mono', monospace !important;
    background: #f1f5f9 !important;
    padding: 1px 5px !important;
    border-radius: 4px !important;
    font-size: 0.87em !important;
}

/* ── METRIC CARDS ── */
[data-testid="metric-container"] {
    background: white !important;
    border-radius: 14px !important;
    padding: 1.25rem 1.5rem !important;
    border: 1px solid #e2e8f0 !important;
    box-shadow: 0 1px 3px rgba(15,23,42,0.05), 0 1px 2px rgba(15,23,42,0.04) !important;
    transition: transform 0.22s cubic-bezier(.4,0,.2,1), box-shadow 0.22s !important;
}
[data-testid="metric-container"]:hover {
    transform: translateY(-3px) !important;
    box-shadow: 0 10px 28px rgba(15,23,42,0.1), 0 0 0 1px rgba(99,102,241,0.18) !important;
}
[data-testid="stMetricLabel"] {
    font-size: 0.73rem !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.07em !important;
    color: #64748b !important;
}
[data-testid="stMetricValue"] {
    font-family: 'Sora', sans-serif !important;
    font-size: 1.95rem !important;
    font-weight: 700 !important;
    color: #0f172a !important;
    line-height: 1.15 !important;
}
[data-testid="stMetricDelta"] { font-size: 0.78rem !important; font-weight: 500 !important; }

/* ── PRIMARY BUTTON ── */
button[kind="primary"] {
    background: linear-gradient(135deg, #6366f1 0%, #4f46e5 100%) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    font-weight: 600 !important;
    font-size: 0.9rem !important;
    font-family: 'DM Sans', sans-serif !important;
    box-shadow: 0 4px 14px rgba(99,102,241,0.32) !important;
    transition: all 0.2s !important;
    letter-spacing: 0.01em !important;
}
button[kind="primary"]:hover {
    box-shadow: 0 8px 22px rgba(99,102,241,0.48) !important;
    transform: translateY(-1px) !important;
}
button[kind="secondary"] {
    background: white !important;
    border: 1.5px solid #e2e8f0 !important;
    color: #475569 !important;
    border-radius: 10px !important;
    font-weight: 500 !important;
    font-family: 'DM Sans', sans-serif !important;
    transition: all 0.2s !important;
}
button[kind="secondary"]:hover {
    border-color: #6366f1 !important;
    color: #6366f1 !important;
    background: #f5f3ff !important;
}
button[kind="tertiary"] {
    background: transparent !important;
    border: 1px solid #e2e8f0 !important;
    color: #64748b !important;
    border-radius: 8px !important;
    font-size: 0.85rem !important;
    transition: all 0.2s !important;
}
button[kind="tertiary"]:hover { background: #f8fafc !important; }

/* ── INPUTS ── */
input, textarea, select {
    border-radius: 10px !important;
    border: 1.5px solid #e2e8f0 !important;
    font-family: 'DM Sans', sans-serif !important;
    font-size: 0.93rem !important;
    color: #1e293b !important;
    transition: border-color 0.2s, box-shadow 0.2s !important;
}
input:focus, textarea:focus {
    border-color: #6366f1 !important;
    box-shadow: 0 0 0 3px rgba(99,102,241,0.12) !important;
    outline: none !important;
}
[data-testid="stNumberInput"] input { border-radius: 10px !important; }
[data-testid="stSelectbox"] > div > div { border-radius: 10px !important; border: 1.5px solid #e2e8f0 !important; }
[data-testid="stMultiSelect"] > div > div { border-radius: 10px !important; border: 1.5px solid #e2e8f0 !important; }

/* ── FILE UPLOADER ── */
[data-testid="stFileUploadDropzone"] {
    background: #f8faff !important;
    border: 2px dashed #c7d2fe !important;
    border-radius: 16px !important;
    padding: 2.25rem !important;
    transition: all 0.2s !important;
}
[data-testid="stFileUploadDropzone"]:hover {
    border-color: #6366f1 !important;
    background: #eef2ff !important;
}

/* ── DATA TABLE ── */
[data-testid="stDataFrame"] {
    border-radius: 12px !important;
    border: 1px solid #e2e8f0 !important;
    overflow: hidden !important;
    box-shadow: 0 1px 3px rgba(15,23,42,0.05) !important;
}

/* ── ALERTS ── */
[data-testid="stAlert"] { border-radius: 10px !important; border: none !important; font-family: 'DM Sans', sans-serif !important; }

/* ── CHAT ── */
[data-testid="stChatMessage"] {
    border: 1px solid #e2e8f0 !important;
    border-radius: 14px !important;
    margin-bottom: 0.75rem !important;
    background: white !important;
    box-shadow: 0 1px 3px rgba(15,23,42,0.04) !important;
}
[data-testid="stChatInput"] { border-radius: 12px !important; }
[data-testid="stChatInput"] > div { border-radius: 12px !important; border: 1.5px solid #e2e8f0 !important; }
[data-testid="stChatInput"] > div:focus-within { border-color: #6366f1 !important; }

/* ── TABS ── */
button[data-baseweb="tab"] {
    font-family: 'DM Sans', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.9rem !important;
}
button[data-baseweb="tab"][aria-selected="true"] {
    color: #6366f1 !important;
    border-bottom-color: #6366f1 !important;
}

/* ── EXPANDER ── */
details {
    border: 1px solid #e2e8f0 !important;
    border-radius: 12px !important;
    overflow: hidden !important;
    margin-bottom: 0.75rem !important;
}
summary {
    background: #f8fafc !important;
    padding: 0.85rem 1.25rem !important;
    font-weight: 500 !important;
    font-family: 'DM Sans', sans-serif !important;
    color: #475569 !important;
    cursor: pointer !important;
    font-size: 0.9rem !important;
}

/* ── PROGRESS ── */
[data-testid="stProgressBar"] > div > div {
    background: linear-gradient(90deg, #6366f1, #06b6d4) !important;
    border-radius: 99px !important;
}

/* ── RADIO (sidebar) ── */
[data-testid="stSidebar"] [data-baseweb="radio"] {
    gap: 2px !important;
}
[data-testid="stSidebar"] label > div > div {
    background-color: #6366f1 !important;
    border-color: #6366f1 !important;
}

/* ── SPINNER ── */
.stSpinner > div { border-top-color: #6366f1 !important; }

/* ── DIVIDER ── */
hr { border: none !important; border-top: 1px solid #e2e8f0 !important; margin: 1.5rem 0 !important; }

/* ── SCROLLBAR ── */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: #f1f5f9; border-radius: 99px; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 99px; }
::-webkit-scrollbar-thumb:hover { background: #94a3b8; }
</style>
"""

st.markdown(CSS, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────
_DEFAULTS = {
    "logged_in": False,
    "username": "",
    "results_df": None,
    "messages": [],
    "last_file_name": None,
    "critical_threshold": 60,
    "warning_threshold": 40,
}
for _k, _v in _DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ─────────────────────────────────────────────────────────────
# MODEL LOADING
# ─────────────────────────────────────────────────────────────
@st.cache_resource
def load_all_models():
    tabular_model = joblib.load("mental_health_model.pkl")
    model_cols    = joblib.load("model_columns.pkl")
    nlp_model     = pipeline("sentiment-analysis", model="distilbert-base-uncased-finetuned-sst-2-english")
    return tabular_model, model_cols, nlp_model

try:
    tabular_model, model_cols, nlp_model = load_all_models()
except FileNotFoundError:
    st.error("⚠️ Critical model files (.pkl) are missing. Run `trainingDemo.py` first to generate them.")
    st.stop()

# ─────────────────────────────────────────────────────────────
# CORE RISK ENGINE
# ─────────────────────────────────────────────────────────────
def calculate_risk(
    age: int,
    hours_worked: int,
    feedback_text: str,
    years_at_company: int = 3,
    stress_level: int = 3,
    sleep_quality: int = 3,
    wlb: int = 3,
    anxiety: int = 3,
    overwhelmed: int = 3,
    gender: str = "Male",
    department: str = "Engineering",
    remote_status: str = "Hybrid",
) -> dict:
    user_input = {
        "Age": age,
        "Work_Hours_Per_Week": hours_worked,
        "Years_at_Company": years_at_company,
        "Screening_Stress_Level": stress_level,
        "Screening_Sleep_Quality": sleep_quality,
        "Screening_Work_Life_Balance": wlb,
        "Screening_Anxiety_Frequency": anxiety,
        "Screening_Overwhelmed_Frequency": overwhelmed,
        "Gender": gender,
        "Department": department,
        "Remote_Status": remote_status,
    }
    input_df = pd.DataFrame([user_input])
    input_df = pd.get_dummies(input_df)
    input_df = input_df.reindex(columns=model_cols, fill_value=0)
    tabular_prob = tabular_model.predict_proba(input_df)[0][1]

    text = str(feedback_text).strip()
    if not text or text.lower() in ("nan", "none", ""):
        nlp_result = {"label": "POSITIVE", "score": 0.5}
    else:
        nlp_result = nlp_model(text[:512])[0]

    if nlp_result["label"] == "NEGATIVE":
        unified = (tabular_prob * 0.4) + (nlp_result["score"] * 0.6)
    else:
        unified = (tabular_prob * 0.5) + ((1 - nlp_result["score"]) * 0.5)

    score = round(unified * 100, 1)
    crit  = st.session_state.critical_threshold
    warn  = st.session_state.warning_threshold

    if score >= crit:
        status = "Critical"
    elif score >= warn:
        status = "Warning"
    else:
        status = "Stable"

    return {
        "Risk_Score":       score,
        "Tabular_Prob":     round(tabular_prob * 100, 1),
        "Sentiment":        nlp_result["label"],
        "Sentiment_Score":  round(nlp_result["score"] * 100, 1),
        "Status":           status,
    }

# ─────────────────────────────────────────────────────────────
# LANGCHAIN AGENT
# ─────────────────────────────────────────────────────────────
@tool
def calculate_burnout_risk_tool(age: int, hours_worked: int, feedback_text: str) -> str:
    """Calculates employee burnout risk given age, weekly hours worked, and free-text feedback."""
    r = calculate_risk(age, hours_worked, feedback_text)
    lines = [
        f"Risk Score: {r['Risk_Score']:.1f}% | Status: {r['Status']}",
        f"Tabular Model: {r['Tabular_Prob']:.1f}% | NLP Sentiment: {r['Sentiment']} ({r['Sentiment_Score']:.0f}% confidence)",
    ]
    if r["Status"] == "Critical":
        lines.append("⚠️ Immediate HR intervention recommended.")
    elif r["Status"] == "Warning":
        lines.append("ℹ️ Schedule a check-in within the next two weeks.")
    return " | ".join(lines)


@st.cache_resource
def get_agent():
    llm   = ChatGoogleGenerativeAI(model="gemini-2.5-flash", temperature=0)
    tools = [calculate_burnout_risk_tool]
    mem   = ConversationBufferMemory(memory_key="chat_history", return_messages=True)
    prompt = ChatPromptTemplate.from_messages([
        ("system", (
            "You are NexaHR's enterprise AI assistant specialising in workforce wellbeing. "
            "Always use the calculate_burnout_risk_tool when evaluating a specific employee. "
            "Structure your responses with: (1) Risk Assessment headline, (2) Score breakdown, "
            "(3) Evidence-based reasoning, (4) Concrete, actionable HR recommendations. "
            "Be concise, professional, and empathetic. Use markdown for clarity."
        )),
        ("placeholder", "{chat_history}"),
        ("human", "{input}"),
        ("placeholder", "{agent_scratchpad}"),
    ])
    agent = create_tool_calling_agent(llm, tools, prompt)
    return AgentExecutor(agent=agent, tools=tools, memory=mem, verbose=False)

# Store the agent inside the user's specific session state!
if "agent_executor" not in st.session_state:
    st.session_state.agent_executor = get_agent()


agent_executor = get_agent()

# ─────────────────────────────────────────────────────────────
# HELPER UI COMPONENTS
# ─────────────────────────────────────────────────────────────
_STATUS_STYLES = {
    "Critical": ("bg:#fff1f2;color:#f43f5e;border:1.5px solid #fecdd3", "🔴"),
    "Warning":  ("bg:#fffbeb;color:#d97706;border:1.5px solid #fde68a", "🟡"),
    "Stable":   ("bg:#f0fdf4;color:#16a34a;border:1.5px solid #bbf7d0", "🟢"),
}

def status_badge(status: str) -> str:
    style, icon = _STATUS_STYLES.get(status, ("bg:#f1f5f9;color:#64748b;border:1px solid #e2e8f0", "⚪"))
    return (
        f'<span style="display:inline-block;padding:3px 12px;border-radius:99px;'
        f'font-size:0.78rem;font-weight:700;letter-spacing:0.03em;{style}">'
        f'{icon} {status}</span>'
    )


def risk_gauge_html(score: float) -> str:
    """SVG circular gauge for the risk score."""
    crit = st.session_state.critical_threshold
    warn = st.session_state.warning_threshold
    if score >= crit:
        color, bg = "#f43f5e", "#fff1f2"
        label = "CRITICAL"
    elif score >= warn:
        color, bg = "#f59e0b", "#fffbeb"
        label = "WARNING"
    else:
        color, bg = "#10b981", "#f0fdf4"
        label = "STABLE"

    r       = 48
    circ    = 2 * 3.14159265 * r
    pct     = min(score / 100.0, 1.0)
    dash    = circ * pct
    gap     = circ - dash

    return f"""
    <div style="text-align:center;padding:1.25rem 0 0.75rem;">
      <svg width="148" height="148" viewBox="0 0 120 120">
        <circle cx="60" cy="60" r="{r}" fill="{bg}" stroke="#e2e8f0" stroke-width="9"/>
        <circle cx="60" cy="60" r="{r}" fill="none" stroke="{color}" stroke-width="9"
          stroke-dasharray="{dash:.2f} {gap:.2f}" stroke-linecap="round"
          transform="rotate(-90 60 60)"/>
        <text x="60" y="54" text-anchor="middle" font-family="Sora,sans-serif"
          font-size="20" font-weight="800" fill="{color}">{score:.0f}%</text>
        <text x="60" y="68" text-anchor="middle" font-family="DM Sans,sans-serif"
          font-size="8" fill="#94a3b8" font-weight="500" letter-spacing="1">RISK SCORE</text>
        <text x="60" y="80" text-anchor="middle" font-family="Sora,sans-serif"
          font-size="8" font-weight="700" fill="{color}" letter-spacing="0.5">{label}</text>
      </svg>
    </div>"""


def mini_bar_html(label: str, value: float, max_val: float = 100.0, color: str = "#6366f1") -> str:
    pct = min(value / max_val * 100, 100)
    return f"""
    <div style="margin-bottom:0.65rem;">
      <div style="display:flex;justify-content:space-between;margin-bottom:4px;">
        <span style="font-size:0.8rem;color:#64748b;font-weight:500;">{label}</span>
        <span style="font-size:0.8rem;color:#1e293b;font-weight:700;font-family:'DM Mono',monospace;">{value:.1f}%</span>
      </div>
      <div style="background:#f1f5f9;border-radius:99px;height:7px;">
        <div style="width:{pct:.1f}%;height:100%;background:{color};border-radius:99px;transition:width 0.5s ease;"></div>
      </div>
    </div>"""


def create_excel_report(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Risk Analysis")
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            ws = writer.sheets["Risk Analysis"]
            hdr_fill = PatternFill(start_color="0B1120", end_color="0B1120", fill_type="solid")
            hdr_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            status_col_idx = None
            for i, col in enumerate(df.columns, 1):
                if col == "Status":
                    status_col_idx = i
            row_fills = {
                "Critical": PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid"),
                "Warning":  PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid"),
                "Stable":   PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid"),
            }
            if status_col_idx:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    status_val = row[status_col_idx - 1].value
                    if status_val in row_fills:
                        for cell in row:
                            cell.fill = row_fills[status_val]
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 38)
        except Exception:
            pass
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# LOGIN PAGE
# ─────────────────────────────────────────────────────────────
if not st.session_state.logged_in:
    # Centered login card
    _, mid, _ = st.columns([1, 1.4, 1])
    with mid:
        st.markdown("""
        <div style="text-align:center;padding:3rem 0 2rem;">
          <div style="width:68px;height:68px;background:linear-gradient(135deg,#6366f1,#06b6d4);
               border-radius:20px;display:inline-flex;align-items:center;justify-content:center;
               font-size:2rem;box-shadow:0 10px 30px rgba(99,102,241,0.4);margin-bottom:1.25rem;">🏢</div>
          <h1 style="font-size:2.2rem !important;margin:0 0 0.35rem !important;">NexaHR</h1>
          <p style="color:#64748b;margin:0;font-size:1rem;">Workforce Intelligence Platform</p>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div style="background:white;border-radius:20px;padding:2.5rem 2.25rem;
             border:1px solid #e2e8f0;box-shadow:0 12px 48px rgba(15,23,42,0.09);">
        """, unsafe_allow_html=True)

        st.markdown("<h3 style='margin-top:0;'>🔐 Sign in</h3>", unsafe_allow_html=True)
        st.markdown("<p style='color:#64748b;font-size:0.9rem;margin-top:-0.5rem;'>Enter your credentials to access the platform</p>", unsafe_allow_html=True)

        with st.form("login_form", clear_on_submit=False):
            uname = st.text_input("Username", placeholder="e.g. admin")
            pwd   = st.text_input("Password", type="password", placeholder="••••••••")
            submitted = st.form_submit_button("Sign In →", use_container_width=True, type="primary")
            if submitted:
                if uname == "admin" and pwd:
                    st.session_state.logged_in = True
                    st.session_state.username  = uname
                    st.rerun()
                else:
                    st.error("Invalid credentials — try username **admin** with any password.")

        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown("""
        <p style="text-align:center;color:#94a3b8;font-size:0.8rem;margin-top:2rem;">
          © 2025 NexaHR · AI-Powered Workforce Intelligence
        </p>""", unsafe_allow_html=True)
    st.stop()


# ─────────────────────────────────────────────────────────────
# AUTHENTICATED APP
# ─────────────────────────────────────────────────────────────

# ── SIDEBAR ──────────────────────────────────────────────────
with st.sidebar:
    # Logo + brand
    st.markdown(f"""
    <div style="padding:1.25rem 1rem 0.5rem;">
      <div style="display:flex;align-items:center;gap:0.75rem;margin-bottom:0.25rem;">
        <div style="width:38px;height:38px;background:linear-gradient(135deg,#6366f1,#06b6d4);
             border-radius:11px;display:flex;align-items:center;justify-content:center;font-size:1.15rem;
             box-shadow:0 4px 12px rgba(99,102,241,0.4);">🏢</div>
        <div>
          <p style="margin:0;font-family:'Sora',sans-serif;font-weight:700;color:#f1f5f9;font-size:1.1rem;line-height:1;">NexaHR</p>
          <p style="margin:0;color:#475569;font-size:0.7rem;letter-spacing:0.04em;">WORKFORCE INTELLIGENCE</p>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    # User pill
    st.markdown(f"""
    <div style="background:#1a2540;border-radius:10px;padding:0.7rem 0.9rem;margin-bottom:0.75rem;border:1px solid #2d3f5c;">
      <p style="margin:0;color:#475569;font-size:0.7rem;text-transform:uppercase;letter-spacing:0.06em;">Signed in as</p>
      <p style="margin:0.2rem 0 0;color:#e2e8f0;font-weight:600;font-size:0.9rem;">👤 {st.session_state.username.capitalize()}</p>
    </div>
    """, unsafe_allow_html=True)

    # Navigation
    st.markdown("<p style='font-size:0.68rem;text-transform:uppercase;letter-spacing:0.09em;color:#334155;margin:0 0 0.4rem;'>Navigation</p>", unsafe_allow_html=True)
    nav = st.radio(
        "nav",
        ["📊  Dashboard", "🔍  Individual Analysis", "🤖  AI Co-Pilot"],
        label_visibility="collapsed",
    )

    st.divider()

    # Risk threshold settings
    with st.expander("⚙️  Risk Thresholds", expanded=False):
        st.session_state.critical_threshold = st.slider(
            "Critical (≥%)", 40, 90, st.session_state.critical_threshold, 5
        )
        st.session_state.warning_threshold = st.slider(
            "Warning (≥%)", 20, 60, st.session_state.warning_threshold, 5
        )

    # Live dataset summary
    if st.session_state.results_df is not None:
        res = st.session_state.results_df
        crit_n = (res["Status"] == "Critical").sum()
        warn_n = (res["Status"] == "Warning").sum()
        stab_n = (res["Status"] == "Stable").sum()
        st.markdown(f"""
        <div style="background:#1a2540;border-radius:10px;padding:0.85rem 0.9rem;margin-top:0.5rem;border:1px solid #2d3f5c;">
          <p style="margin:0 0 0.6rem;color:#475569;font-size:0.68rem;text-transform:uppercase;letter-spacing:0.06em;">Loaded Dataset</p>
          <p style="margin:0 0 0.2rem;color:#94a3b8;font-size:0.78rem;font-family:'DM Mono',monospace;">
            📁 {str(st.session_state.last_file_name)[:22]}…</p>
          <p style="margin:0 0 0.15rem;color:#f87171;font-size:0.8rem;">🔴 Critical: <strong style="color:#f87171">{crit_n}</strong> ({crit_n/len(res)*100:.0f}%)</p>
          <p style="margin:0 0 0.15rem;color:#fbbf24;font-size:0.8rem;">🟡 Warning: <strong style="color:#fbbf24">{warn_n}</strong> ({warn_n/len(res)*100:.0f}%)</p>
          <p style="margin:0;color:#34d399;font-size:0.8rem;">🟢 Stable: <strong style="color:#34d399">{stab_n}</strong> ({stab_n/len(res)*100:.0f}%)</p>
        </div>
        """, unsafe_allow_html=True)

    st.divider()
    if st.button("🚪  Sign Out", use_container_width=True):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()


# ═════════════════════════════════════════════════════════════
# VIEW 1 — DASHBOARD
# ═════════════════════════════════════════════════════════════
if nav == "📊  Dashboard":

    st.markdown("""
    <div style="margin-bottom:1.5rem;">
      <h1>📊 Enterprise Dashboard</h1>
      <p style="color:#64748b;margin:0.3rem 0 0;font-size:1rem;">
        Upload workforce survey data for AI-driven burnout & sentiment analysis
      </p>
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader(
        "Drop HR survey CSV here, or click to browse",
        type="csv",
        help="CSV should include: Age, Work_Hours_Per_Week, Department, and Feedback / Feedback_Text columns.",
    )

    if uploaded is not None:
        df_raw = pd.read_csv(uploaded)
        st.session_state.last_file_name = uploaded.name

        with st.expander(f"📋 Preview — {uploaded.name}  ({len(df_raw):,} rows)", expanded=False):
            st.dataframe(df_raw.head(8), use_container_width=True, hide_index=True)

        if st.button("⚡  Run Workforce Analysis", type="primary", key="run_batch"):
            prog  = st.progress(0, text="Initialising AI engine…")
            total = len(df_raw)
            rows  = []
            for i, (_, row) in enumerate(df_raw.iterrows()):
                age      = int(row.get("Age", 30))
                hours    = int(row.get("Work_Hours_Per_Week", 40))
                feedback = str(row.get("Feedback_Text", row.get("Feedback", "")))
                calc     = calculate_risk(age, hours, feedback)
                d = row.to_dict()
                d.update(calc)
                rows.append(d)
                prog.progress((i + 1) / total, text=f"Processing {i+1} / {total} employees…")
            prog.empty()
            st.session_state.results_df = pd.DataFrame(rows)
            st.rerun()

    # ── Results ──────────────────────────────────────────────
    if st.session_state.results_df is not None:
        res    = st.session_state.results_df
        crit_n = int((res["Status"] == "Critical").sum())
        warn_n = int((res["Status"] == "Warning").sum())
        stab_n = int((res["Status"] == "Stable").sum())
        avg_r  = res["Risk_Score"].mean()
        neg_n  = int((res["Sentiment"] == "NEGATIVE").sum())
        total  = len(res)

        st.markdown("---")

        # KPI cards
        k1, k2, k3, k4, k5 = st.columns(5)
        with k1:
            st.metric("👥 Total Analysed",    total)
        with k2:
            st.metric("🔴 Critical Risk",     crit_n,
                      f"{crit_n/total*100:.1f}% of workforce", delta_color="inverse")
        with k3:
            st.metric("🟡 At Warning",        warn_n,
                      f"{warn_n/total*100:.1f}%", delta_color="off")
        with k4:
            st.metric("📊 Avg Risk Score",    f"{avg_r:.1f}%")
        with k5:
            st.metric("💬 Negative Sentiment", neg_n,
                      f"{neg_n/total*100:.1f}%", delta_color="inverse")

        st.markdown("---")

        # ── Charts row ───────────────────────────────────────
        if PLOTLY:
            c1, c2, c3 = st.columns([1, 1.15, 1.25])

            with c1:
                fig = go.Figure(go.Pie(
                    labels=["Critical", "Warning", "Stable"],
                    values=[crit_n, warn_n, stab_n],
                    hole=0.66,
                    marker_colors=["#f43f5e", "#f59e0b", "#10b981"],
                    textinfo="percent",
                    textfont_size=12,
                    hovertemplate="%{label}: %{value} employees<extra></extra>",
                ))
                fig.update_layout(
                    title=dict(text="Risk Distribution", font=dict(family="Sora", size=14, color="#1e293b"), x=0.5),
                    showlegend=True,
                    legend=dict(orientation="h", yanchor="bottom", y=-0.22,
                                xanchor="center", x=0.5, font=dict(family="DM Sans", size=11)),
                    margin=dict(t=48, b=10, l=10, r=10), height=290,
                    paper_bgcolor="white", plot_bgcolor="white",
                    annotations=[dict(text=f"<b>{total}</b><br><span style='font-size:10px'>total</span>",
                                      x=0.5, y=0.5, showarrow=False,
                                      font=dict(family="Sora", size=15, color="#0f172a"))]
                )
                st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

            with c2:
                if "Department" in res.columns:
                    dept_risk = res.groupby("Department")["Risk_Score"].mean().sort_values()
                    fig2 = px.bar(
                        x=dept_risk.values, y=dept_risk.index, orientation="h",
                        title="Avg Risk by Department",
                        labels={"x": "Avg Risk (%)", "y": ""},
                        color=dept_risk.values,
                        color_continuous_scale=[[0,"#10b981"],[0.4,"#f59e0b"],[1,"#f43f5e"]],
                        color_continuous_midpoint=st.session_state.warning_threshold,
                    )
                    fig2.update_layout(
                        title=dict(font=dict(family="Sora", size=14, color="#1e293b"), x=0.5),
                        coloraxis_showscale=False,
                        margin=dict(t=48, b=10, l=10, r=10), height=290,
                        paper_bgcolor="white", plot_bgcolor="white",
                        font=dict(family="DM Sans"),
                        xaxis=dict(gridcolor="#f1f5f9"),
                        yaxis=dict(gridcolor="#f1f5f9"),
                    )
                    st.plotly_chart(fig2, use_container_width=True, config={"displayModeBar": False})
                else:
                    st.info("Add a **Department** column to see the department breakdown chart.")

            with c3:
                if "Work_Hours_Per_Week" in res.columns:
                    hover_data = {k: True for k in ["Department", "Age", "Sentiment"]
                                  if k in res.columns}
                    hover_data["Risk_Score"] = True
                    fig3 = px.scatter(
                        res, x="Work_Hours_Per_Week", y="Risk_Score",
                        color="Status",
                        color_discrete_map={"Critical": "#f43f5e", "Warning": "#f59e0b", "Stable": "#10b981"},
                        title="Work Hours vs Risk Score",
                        labels={"Work_Hours_Per_Week": "Hours / Week", "Risk_Score": "Risk Score (%)"},
                        opacity=0.75,
                        hover_data=hover_data,
                    )
                    fig3.update_traces(marker=dict(size=7, line=dict(width=0.5, color="white")))
                    fig3.update_layout(
                        title=dict(font=dict(family="Sora", size=14, color="#1e293b"), x=0.5),
                        margin=dict(t=48, b=10, l=10, r=10), height=290,
                        paper_bgcolor="white", plot_bgcolor="white",
                        font=dict(family="DM Sans"),
                        legend=dict(font=dict(family="DM Sans", size=10),
                                    orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5),
                        xaxis=dict(gridcolor="#f1f5f9"),
                        yaxis=dict(gridcolor="#f1f5f9"),
                    )
                    st.plotly_chart(fig3, use_container_width=True, config={"displayModeBar": False})
        else:
            # Fallback: native Streamlit charts
            col_a, col_b = st.columns(2)
            with col_a:
                st.markdown("**Risk Distribution**")
                st.bar_chart({"Critical": crit_n, "Warning": warn_n, "Stable": stab_n})
            with col_b:
                if "Department" in res.columns:
                    st.markdown("**Avg Risk by Department**")
                    st.bar_chart(res.groupby("Department")["Risk_Score"].mean())

        st.markdown("---")

        # ── Critical Alerts Panel ─────────────────────────────
        if crit_n > 0:
            st.markdown("#### 🚨 Critical Risk Alerts")
            top_critical = res[res["Status"] == "Critical"].nlargest(min(5, crit_n), "Risk_Score")
            alert_cols   = st.columns(len(top_critical))
            for col, (_, emp) in zip(alert_cols, top_critical.iterrows()):
                emp_id = emp.get("Employee_ID", "—")
                dept   = emp.get("Department", "—")
                score  = emp["Risk_Score"]
                sent   = emp.get("Sentiment", "—")
                with col:
                    st.markdown(f"""
                    <div style="background:white;border-radius:14px;padding:1rem 0.75rem;
                         border:1.5px solid #fecdd3;text-align:center;
                         box-shadow:0 4px 16px rgba(244,63,94,0.1);">
                      <p style="margin:0;font-size:0.72rem;color:#94a3b8;letter-spacing:0.04em;text-transform:uppercase;">{dept}</p>
                      <p style="margin:0.3rem 0;font-family:'Sora',sans-serif;font-weight:700;
                           color:#0f172a;font-size:0.88rem;">{emp_id}</p>
                      <p style="margin:0;font-size:1.75rem;font-weight:800;color:#f43f5e;
                           font-family:'Sora',sans-serif;line-height:1;">{score:.0f}%</p>
                      <p style="margin:0.2rem 0 0;font-size:0.68rem;color:#f43f5e;
                           font-weight:700;letter-spacing:0.05em;">CRITICAL · {sent}</p>
                    </div>
                    """, unsafe_allow_html=True)
            st.markdown("<div style='height:1rem'></div>", unsafe_allow_html=True)

        # ── Filter + Table ────────────────────────────────────
        st.markdown("#### 📋 Employee Ledger")

        f1, f2, f3, f4 = st.columns([1.4, 1.2, 1.1, 1.1])
        with f1:
            status_filter = st.multiselect(
                "Status", ["Critical", "Warning", "Stable"],
                default=["Critical", "Warning", "Stable"],
                label_visibility="collapsed", placeholder="Filter by status",
            )
        with f2:
            dept_options = sorted(res["Department"].dropna().unique()) if "Department" in res.columns else []
            dept_filter  = st.multiselect("Dept", dept_options, label_visibility="collapsed",
                                           placeholder="All departments")
        with f3:
            sort_options = [c for c in ["Risk_Score", "Age", "Work_Hours_Per_Week"] if c in res.columns]
            sort_by = st.selectbox("Sort by", sort_options, label_visibility="collapsed")
        with f4:
            sent_filter = st.selectbox("Sentiment", ["All", "POSITIVE", "NEGATIVE"],
                                        label_visibility="collapsed")

        mask = res["Status"].isin(status_filter or ["Critical", "Warning", "Stable"])
        if dept_filter:
            mask &= res["Department"].isin(dept_filter)
        if sent_filter != "All":
            mask &= res["Sentiment"] == sent_filter
        filtered = res[mask].sort_values(sort_by, ascending=False)

        disp_cols = ["Employee_ID", "Department", "Age", "Work_Hours_Per_Week",
                     "Risk_Score", "Tabular_Prob", "Sentiment", "Status"]
        disp_cols = [c for c in disp_cols if c in filtered.columns]

        st.caption(f"Showing {len(filtered):,} of {total:,} employees")
        st.dataframe(filtered[disp_cols], use_container_width=True, height=420, hide_index=True)

        dl1, dl2, _ = st.columns([1, 1, 3])
        with dl1:
            st.download_button(
                "📥 Download CSV",
                filtered.to_csv(index=False).encode("utf-8"),
                f"nexahr_{st.session_state.last_file_name or 'report'}.csv",
                "text/csv", use_container_width=True,
            )
        with dl2:
            st.download_button(
                "📊 Download Excel",
                create_excel_report(filtered),
                f"nexahr_{st.session_state.last_file_name or 'report'}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


# ═════════════════════════════════════════════════════════════
# VIEW 2 — INDIVIDUAL ANALYSIS (NEW)
# ═════════════════════════════════════════════════════════════
elif nav == "🔍  Individual Analysis":

    st.markdown("""
    <div style="margin-bottom:1.5rem;">
      <h1>🔍 Individual Employee Analysis</h1>
      <p style="color:#64748b;margin:0.3rem 0 0;font-size:1rem;">
        Deep-dive assessment using all screening dimensions plus NLP text analysis
      </p>
    </div>
    """, unsafe_allow_html=True)

    form_col, result_col = st.columns([1.5, 1], gap="large")

    with form_col:
        st.markdown("""<div style="background:white;border-radius:16px;padding:1.75rem 1.75rem 1.5rem;
                    border:1px solid #e2e8f0;box-shadow:0 1px 4px rgba(15,23,42,0.05);">""",
                    unsafe_allow_html=True)

        st.markdown("#### 👤 Employee Profile")

        r1c1, r1c2, r1c3 = st.columns(3)
        with r1c1:
            age    = st.number_input("Age", 18, 65, 32,   key="ia_age")
            gender = st.selectbox("Gender", ["Male", "Female", "Non-binary"], key="ia_gender")
        with r1c2:
            years  = st.number_input("Years at Company", 0, 40, 3, key="ia_years")
            dept   = st.selectbox("Department",
                                  ["Engineering","Sales","HR","Finance","Marketing","Operations"],
                                  key="ia_dept")
        with r1c3:
            hours  = st.number_input("Work Hours / Week", 20, 84, 45, key="ia_hours")
            remote = st.selectbox("Work Type",
                                  ["Hybrid","On-site","Full-time Remote"],
                                  key="ia_remote")

        st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
        st.markdown("#### 📋 Screening Scores")
        st.caption("Rate each dimension: **1 = Very Low / Excellent** · **5 = Very High / Poor**")

        sc1, sc2 = st.columns(2)
        with sc1:
            stress      = st.slider("😰 Stress Level",          1, 5, 3, key="ia_stress")
            sleep       = st.slider("😴 Sleep Quality",         1, 5, 3, key="ia_sleep",
                                     help="1 = Poor sleep · 5 = Excellent sleep")
            wlb         = st.slider("⚖️ Work-Life Balance",     1, 5, 3, key="ia_wlb",
                                     help="1 = Poor balance · 5 = Excellent balance")
        with sc2:
            anxiety     = st.slider("😟 Anxiety Frequency",    1, 5, 3, key="ia_anxiety")
            overwhelmed = st.slider("🌀 Overwhelmed Frequency", 1, 5, 3, key="ia_overwhelmed")

        st.markdown("#### 💬 Written Feedback")
        feedback = st.text_area(
            "Paste employee comments, interview notes, or survey text:",
            height=110, key="ia_feedback",
            placeholder="e.g. 'I've been struggling lately. The deadlines are unrealistic…'",
        )

        run_analysis = st.button("⚡  Analyse Employee", type="primary",
                                  use_container_width=True, key="ia_run")

        st.markdown("</div>", unsafe_allow_html=True)

    # ── Results panel ─────────────────────────────────────────
    with result_col:
        if run_analysis:
            with st.spinner("Running multi-modal analysis…"):
                result = calculate_risk(
                    age, hours, feedback,
                    years, stress, sleep, wlb, anxiety, overwhelmed,
                    gender, dept, remote,
                )

            # Gauge
            st.markdown(risk_gauge_html(result["Risk_Score"]), unsafe_allow_html=True)
            st.markdown(
                f'<div style="text-align:center;margin-bottom:1rem;">'
                f'{status_badge(result["Status"])}</div>',
                unsafe_allow_html=True,
            )

            # Score breakdown bars
            st.markdown("""<div style="background:white;border-radius:14px;padding:1.25rem 1.25rem 1rem;
                        border:1px solid #e2e8f0;margin-bottom:0.75rem;">
                        <p style='font-size:0.78rem;font-weight:600;text-transform:uppercase;
                        letter-spacing:0.06em;color:#64748b;margin:0 0 0.9rem;'>Score Breakdown</p>
                        """, unsafe_allow_html=True)
            st.markdown(mini_bar_html("Tabular ML Risk",   result["Tabular_Prob"],  color="#6366f1"),
                        unsafe_allow_html=True)
            st.markdown(mini_bar_html("NLP Confidence",    result["Sentiment_Score"], color="#06b6d4"),
                        unsafe_allow_html=True)
            st.markdown(mini_bar_html("Unified Risk Score", result["Risk_Score"],   color=(
                "#f43f5e" if result["Status"]=="Critical"
                else "#f59e0b" if result["Status"]=="Warning"
                else "#10b981"
            )), unsafe_allow_html=True)

            sent_color = "#f43f5e" if result["Sentiment"] == "NEGATIVE" else "#10b981"
            st.markdown(
                f'<p style="font-size:0.82rem;color:#64748b;margin:0.25rem 0 0;">'
                f'NLP detected: <strong style="color:{sent_color};">{result["Sentiment"]}</strong> sentiment</p>',
                unsafe_allow_html=True,
            )
            st.markdown("</div>", unsafe_allow_html=True)

            # Recommendation
            if result["Status"] == "Critical":
                st.error(
                    "🚨 **Immediate Action Required**\n\n"
                    "Schedule an urgent 1:1 with HR. Consider referring to an Employee "
                    "Assistance Programme (EAP). Conduct an urgent workload audit."
                )
            elif result["Status"] == "Warning":
                st.warning(
                    "⚠️ **Monitor Closely**\n\n"
                    "Schedule a wellbeing check-in within two weeks. Review workload "
                    "distribution and leadership support structures."
                )
            else:
                st.success(
                    "✅ **Stable Profile**\n\n"
                    "Employee appears well-supported. Maintain current engagement "
                    "cadence with periodic pulse checks."
                )
        else:
            st.markdown("""
            <div style="background:white;border-radius:16px;padding:2.5rem 1.5rem;
                 border:1px solid #e2e8f0;text-align:center;margin-top:0.5rem;">
              <p style="font-size:3rem;margin:0 0 0.75rem;">🔍</p>
              <p style="color:#94a3b8;margin:0;font-size:0.95rem;line-height:1.6;">
                Complete the employee profile on the left<br>and click
                <strong style="color:#6366f1;">Analyse Employee</strong>
                to generate an instant risk assessment.
              </p>
            </div>
            """, unsafe_allow_html=True)


# ═════════════════════════════════════════════════════════════
# VIEW 3 — AI CO-PILOT
# ═════════════════════════════════════════════════════════════
elif nav == "🤖  AI Co-Pilot":

    st.markdown("""
    <div style="margin-bottom:1rem;">
      <h1>🤖 HR AI Co-Pilot</h1>
      <p style="color:#64748b;margin:0.3rem 0 0;font-size:1rem;">
        Powered by <strong>Google Gemini 2.5 Flash</strong> · Conversational workforce risk analysis
      </p>
    </div>
    """, unsafe_allow_html=True)

    # How-to banner
    st.markdown("""
    <div style="background:linear-gradient(135deg,#eef2ff 0%,#f0f9ff 100%);
         border-left:4px solid #6366f1;border-radius:0 10px 10px 0;
         padding:0.9rem 1.2rem;margin-bottom:1rem;">
      <p style="margin:0;font-weight:600;color:#3730a3;font-size:0.9rem;">💬 How to use</p>
      <p style="margin:0.3rem 0 0;color:#4338ca;font-size:0.85rem;">
        Describe an employee's situation — their <strong>age</strong>, <strong>weekly hours</strong>,
        and any <strong>feedback or concerns</strong>. The AI will calculate a burnout risk score and
        provide structured, evidence-based HR recommendations.
      </p>
    </div>
    """, unsafe_allow_html=True)

    # Quick templates
    st.markdown("<p style='font-size:0.8rem;font-weight:600;color:#64748b;text-transform:uppercase;letter-spacing:0.05em;margin-bottom:0.5rem;'>Quick Templates</p>", unsafe_allow_html=True)

    TEMPLATES = [
        ("🔥 Critical Burnout", 42,  62, "I'm completely burned out. The deadlines are impossible and I haven't slept properly in weeks. I'm close to quitting."),
        ("✅ Healthy Check",    29,  38, "Work is going well. Good team balance, my manager is supportive and deadlines are realistic."),
        ("⚠️ Borderline Case", 44,  54, "Some pressure lately but managing. Would benefit from more leadership communication and clearer priorities."),
        ("🆕 New Hire",        24,  43, "Still finding my feet. The onboarding was a bit overwhelming but I'm excited about the role."),
    ]

    tmpl_cols = st.columns(4)
    for col, (label, age_t, hrs_t, fbk_t) in zip(tmpl_cols, TEMPLATES):
        with col:
            if st.button(label, key=f"t_{label}", use_container_width=True):
                prompt_t = (
                    f"Analyse this employee: Age {age_t}, working {hrs_t} hours per week. "
                    f'Feedback: "{fbk_t}"'
                )
                st.session_state.messages.append({"role": "user", "content": prompt_t})
                with st.spinner("Analysing…"):
                    try:
                        resp = st.session_state.agent_executor.invoke({"input": prompt_t})
                        reply = resp["output"]
                    except Exception as e:
                        reply = f"⚠️ Analysis error: {e}"
                st.session_state.messages.append({"role": "assistant", "content": reply})
                st.rerun()

    st.markdown("---")

    # Controls row
    ctl1, ctl2, ctl3, _ = st.columns([1, 1, 1, 3])
    with ctl1:
        if st.button("🗑️  Clear Chat", use_container_width=True):
            st.session_state.messages = []
            st.rerun()
    with ctl2:
        if st.session_state.messages:
            transcript = "\n\n".join(
                f"[{m['role'].upper()}]\n{m['content']}"
                for m in st.session_state.messages
            )
            st.download_button(
                "📥  Export Chat", transcript.encode(), "nexahr_copilot_session.txt",
                "text/plain", use_container_width=True,
            )
    with ctl3:
        msg_count = len([m for m in st.session_state.messages if m["role"] == "user"])
        st.markdown(
            f'<p style="font-size:0.8rem;color:#94a3b8;padding:0.5rem 0;margin:0;">'
            f'{msg_count} message{"s" if msg_count!=1 else ""} sent this session</p>',
            unsafe_allow_html=True,
        )

    # Initialise conversation
    if not st.session_state.messages:
        st.session_state.messages = [{
            "role": "assistant",
            "content": (
                "👋 Hello! I'm your **NexaHR AI Co-Pilot**.\n\n"
                "Describe an employee's situation — their age, weekly hours, and any feedback "
                "or concerns — and I'll provide a detailed burnout risk analysis with actionable recommendations.\n\n"
                "**Tip:** Use the Quick Templates above to get started instantly."
            ),
        }]

    # Chat history
    for msg in st.session_state.messages:
        icon = "🤖" if msg["role"] == "assistant" else "👤"
        with st.chat_message(msg["role"], avatar=icon):
            st.markdown(msg["content"])

    # User input
    if user_input := st.chat_input(
        "Describe an employee case (e.g. 'Sarah, 38, 60 hrs/week, says she's burned out…')"
    ):
        st.session_state.messages.append({"role": "user", "content": user_input})
        with st.chat_message("user", avatar="👤"):
            st.markdown(user_input)

        with st.chat_message("assistant", avatar="🤖"):
            with st.spinner("Analysing employee profile…"):
                try:
                    resp  = st.session_state.agent_executor.invoke({"input": user_input})
                    reply = resp["output"]
                except Exception as e:
                    reply = f"⚠️ Analysis error: {e}"
            st.markdown(reply)
        st.session_state.messages.append({"role": "assistant", "content": reply})